from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from flask import Flask, render_template, request, redirect, url_for
import sqlite3
from openpyxl.styles import Border, Side, Font, Alignment
from datetime import datetime, timedelta
from openpyxl import Workbook
from flask import send_file
import os
from datetime import date
import os

app = Flask(__name__)
from datetime import date


app.secret_key = 'clave_super_secreta_123'

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = "login"

class User(UserMixin):
    def __init__(self, id):
        self.id = id

# Usuario fijo (luego podemos hacerlo con base de datos)
usuarios = {
    "admin": {"password": "1234"}
}

@login_manager.user_loader
def load_user(user_id):
    return User(user_id)


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        if username in usuarios and usuarios[username]['password'] == password:
            user = User(username)
            login_user(user)
            return redirect(url_for('index'))
        else:
            return "Usuario o contraseña incorrectos"

    return render_template('login.html')



@app.route("/dashboard")
@login_required
def dashboard():
    return render_template("inicio.html")

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))



# Crear base de datos si no existe
def init_db():
    conn = sqlite3.connect("prestamos.db")
    cursor = conn.cursor()
    
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS clientes (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nombre TEXT,
        apellidos TEXT,
        dni TEXT,
        direccion TEXT,
        telefono TEXT,
        monto REAL,
        interes REAL,
        total REAL,
        tipo_pago TEXT,
        cuotas INTEGER
    )
    """)
    
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS cronograma (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        cliente_id INTEGER,
        fecha_pago TEXT,
        cuota REAL,
        estado TEXT
    )
    """)
    
    conn.commit()
    conn.close()

init_db()

@app.route("/")
def index():
    conn = sqlite3.connect("prestamos.db")
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM clientes")
    clientes = cursor.fetchall()
    conn.close()
    return render_template("index.html", clientes=clientes)


@app.route("/nuevo", methods=["GET", "POST"])
def nuevo():
    if request.method == "POST":
        nombre = request.form["nombre"]
        apellidos = request.form["apellidos"]
        dni = request.form["dni"]
        direccion = request.form["direccion"]
        telefono = request.form["telefono"]
        monto = float(request.form["monto"])
        interes = float(request.form["interes"])
        tipo_pago = request.form["tipo_pago"]
        cuotas = int(request.form["cuotas"])
        
        total = monto + (monto * interes / 100)
        valor_cuota = round(total / cuotas, 2)
        
        conn = sqlite3.connect("prestamos.db")
        cursor = conn.cursor()
        
        cursor.execute("""
        INSERT INTO clientes (nombre, apellidos, dni, direccion, telefono, monto, interes, total, tipo_pago, cuotas)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (nombre, apellidos, dni, direccion, telefono, monto, interes, total, tipo_pago, cuotas))
        
        cliente_id = cursor.lastrowid
        
        fecha_actual = datetime.today()
        
        for i in range(cuotas):
            if tipo_pago == "semanal":
                fecha_pago = fecha_actual + timedelta(weeks=i+1)
            elif tipo_pago == "quincenal":
                fecha_pago = fecha_actual + timedelta(days=15*(i+1))
            else:
                fecha_pago = fecha_actual + timedelta(days=30*(i+1))
            
            cursor.execute("""
            INSERT INTO cronograma (cliente_id, fecha_pago, cuota, estado)
            VALUES (?, ?, ?, ?)
            """, (cliente_id, fecha_pago.strftime("%d/%m/%Y"), valor_cuota, "Pendiente"))
        
        conn.commit()
        conn.close()
        
        return redirect("/")
    
    return render_template("nuevo.html")


@app.route('/editar_cliente/<int:cliente_id>', methods=['GET', 'POST'])
@login_required
def editar_cliente(cliente_id):
    conn = sqlite3.connect('prestamos.db')
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    # SI ENVÍA FORMULARIO
    if request.method == 'POST':
        nombre = request.form['nombre']
        apellidos = request.form['apellidos']
        dni = request.form['dni']
        direccion = request.form['direccion']
        telefono = request.form['telefono']
        monto = float(request.form['monto'])
        interes = float(request.form['interes'])
        tipo_pago = request.form['tipo_pago']
        cuotas = int(request.form['cuotas'])

        total = monto + (monto * interes / 100)

        cursor.execute("""
            UPDATE clientes
            SET nombre=?, apellidos=?, dni=?, direccion=?, telefono=?,
                monto=?, interes=?, total=?, tipo_pago=?, cuotas=?
            WHERE id=?
        """, (
            nombre, apellidos, dni, direccion, telefono,
            monto, interes, total, tipo_pago, cuotas, cliente_id
        ))

        conn.commit()
        conn.close()
        return redirect('/')

    # SI SOLO ENTRA A VER
    cursor.execute("SELECT * FROM clientes WHERE id=?", (cliente_id,))
    cliente = cursor.fetchone()
    conn.close()

    return render_template('editar_cliente.html', cliente=cliente)



@app.route("/cronograma/<int:cliente_id>")
def ver_cronograma(cliente_id):
    conn = sqlite3.connect("prestamos.db")
    cursor = conn.cursor()

    cursor.execute("SELECT total FROM clientes WHERE id = ?", (cliente_id,))
    cliente = cursor.fetchone()
    total_prestamo = cliente[0]

    cursor.execute("SELECT * FROM cronograma WHERE cliente_id = ?", (cliente_id,))
    cuotas = cursor.fetchall()

    from datetime import datetime

    hoy = datetime.today()

    cuotas_actualizadas = []
    total_pagado = 0
    cuotas_pagadas = 0

for cuota in cuotas:
    fecha_str = cuota[2]

    if "-" in fecha_str:
        fecha_pago = datetime.strptime(fecha_str, "%Y-%m-%d")
    else:
        fecha_pago = datetime.strptime(fecha_str, "%d/%m/%Y")

    estado = cuota[4]


	 monto_original = cuota[3]
		
    dias_retraso = (hoy - fecha_pago).days
    mora = 0
    puntualidad = "-"
	
		    if estado == "Pendiente" and dias_retraso > 0:
		        mora = round(monto_original * 0.02 * dias_retraso, 2)
		        estado = "Vencido"
		        puntualidad = "Retrasado"
		    elif estado == "Pagado" and dias_retraso <= 0:
		        puntualidad = "Puntual"
		
		    monto_final = monto_original + mora
		
		    if estado == "Pagado":
		        total_pagado += monto_original
		        cuotas_pagadas += 1
		
		    cuotas_actualizadas.append((
		        cuota[0],
		        cuota[1],
		        cuota[2],
		        monto_final,
		        estado,
		        mora,
		        puntualidad
		    ))


    total_pendiente = total_prestamo - total_pagado

    conn.close()

    return render_template(
    "cronograma.html",
    cuotas=cuotas_actualizadas,
    total_prestamo=total_prestamo,
    total_pagado=total_pagado,
    total_pendiente=total_pendiente,
    cuotas_pagadas=cuotas_pagadas,
    cliente_id=cliente_id
)

@app.route('/editar_cronograma/<int:cliente_id>', methods=['GET', 'POST'])
@login_required
def editar_cronograma(cliente_id):

    conn = sqlite3.connect('prestamos.db')
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM clientes WHERE id = ?", (cliente_id,))
    cliente = cursor.fetchone()

    cursor.execute("SELECT * FROM cronograma WHERE cliente_id = ?", (cliente_id,))
    cuotas_db = cursor.fetchall()

    # CREAR LISTA NUEVA
    cuotas = []
    from datetime import datetime

    for c in cuotas_db:
        fecha_obj = datetime.strptime(c['fecha_pago'], "%d/%m/%Y")
        cuotas.append({
            'id': c['id'],
            'fecha_pago': fecha_obj.strftime("%Y-%m-%d"),
            'cuota': c['cuota'],
            'estado': c['estado']
        })
	
    if request.method == 'POST':
        for cuota in cuotas:
            cuota_id = cuota['id']

			fecha_html = request.form.get(f'fecha_{cuota_id}')
			nuevo_estado = request.form.get(f'estado_{cuota_id}')
			nuevo_monto = request.form.get(f'monto_{cuota_id}')
			
			# Convertir formato HTML (YYYY-MM-DD) a formato BD (DD/MM/YYYY)
			fecha_obj = datetime.strptime(fecha_html, "%Y-%m-%d")
			nueva_fecha = fecha_obj.strftime("%d/%m/%Y")

            cursor.execute("""
                UPDATE cronograma
                SET fecha_pago = ?, cuota = ?, estado = ?
                WHERE id = ?
            """, (nueva_fecha, nuevo_monto, nuevo_estado, cuota_id))

        conn.commit()
        conn.close()
        return redirect(f'/cronograma/{cliente_id}')

    conn.close()
    return render_template('editar_cronograma.html', cliente=cliente, cuotas=cuotas)




@app.route('/actualizar_bd')
def actualizar_bd():
    conn = sqlite3.connect("prestamos.db")
    cursor = conn.cursor()

    try:
        cursor.execute("ALTER TABLE cronograma ADD COLUMN mora REAL DEFAULT 0")
    except:
        pass

    try:
        cursor.execute("ALTER TABLE cronograma ADD COLUMN puntualidad TEXT DEFAULT '-'")
    except:
        pass

    conn.commit()
    conn.close()

    return "Base de datos actualizada"




@app.context_processor
def contar_vencidos():
    hoy = date.today().strftime("%Y-%m-%d")

    conexion = sqlite3.connect('prestamos.db')
    cursor = conexion.cursor()

    cursor.execute("""
        SELECT COUNT(*)
        FROM cronograma
        WHERE fecha_pago < ?
        AND estado = 'Pendiente'
    """, (hoy,))

    cantidad = cursor.fetchone()[0]

    conexion.close()

    return dict(total_vencidos=cantidad)



@app.route('/eliminar/<int:cliente_id>')
@login_required
def eliminar(cliente_id):

    conexion = sqlite3.connect('prestamos.db')
    cursor = conexion.cursor()

    cursor.execute("DELETE FROM cronograma WHERE cliente_id = ?", (cliente_id,))
    cursor.execute("DELETE FROM clientes WHERE id = ?", (cliente_id,))

    conexion.commit()
    conexion.close()

    return redirect('/reporte')








@app.route('/ver_tablas')
def ver_tablas():
    conexion = sqlite3.connect('prestamos.db')
    cursor = conexion.cursor()

    cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
    tablas = cursor.fetchall()

    conexion.close()

    return str(tablas)



@app.route("/pagar/<int:cuota_id>")
def pagar(cuota_id):
    conn = sqlite3.connect("prestamos.db")
    cursor = conn.cursor()
    cursor.execute("UPDATE cronograma SET estado = 'Pagado' WHERE id = ?", (cuota_id,))
    conn.commit()
    conn.close()
    return redirect("/")

@app.route('/reporte', methods=['GET', 'POST'])
@login_required
def reporte():
    cantidad = None
    resultados = []

    if request.method == 'POST':
        fecha_inicio = request.form['fecha_inicio']

        conexion = sqlite3.connect('prestamos.db')
        conexion.row_factory = sqlite3.Row
        cursor = conexion.cursor()

        cursor.execute("""
   	    SELECT 
      	      cronograma.id AS cronograma_id,
              clientes.id AS cliente_id,
              clientes.nombre,
              clientes.apellidos,
              clientes.telefono,
              clientes.direccion,
              clientes.interes,
              cronograma.fecha_pago,
              cronograma.estado
    	    FROM cronograma
            JOIN clientes ON cronograma.cliente_id = clientes.id
            WHERE cronograma.fecha_pago >= ?
        """, (fecha_inicio,))


        resultados = cursor.fetchall()
        cantidad = len(resultados)

        conexion.close()

    return render_template('reporte.html', cantidad=cantidad, resultados=resultados)





@app.route('/ver_columnas')
def ver_columnas():
    conexion = sqlite3.connect('prestamos.db')
    cursor = conexion.cursor()

    cursor.execute("PRAGMA table_info(cronograma);")
    columnas = cursor.fetchall()

    conexion.close()

    return str(columnas)


@app.route('/vencidos')
@login_required
def vencidos():
    hoy = date.today().strftime("%Y-%m-%d")

    conexion = sqlite3.connect('prestamos.db')
    conexion.row_factory = sqlite3.Row
    cursor = conexion.cursor()

    cursor.execute("""
        SELECT 
            cronograma.id AS cronograma_id,
            clientes.nombre,
            clientes.apellidos,
            cronograma.fecha_pago,
            cronograma.estado
        FROM cronograma
        INNER JOIN clientes 
            ON cronograma.cliente_id = clientes.id
        WHERE cronograma.fecha_pago < ?
        AND cronograma.estado = 'Pendiente'
    """, (hoy,))

    resultados = cursor.fetchall()
    cantidad = len(resultados)

    conexion.close()

    return render_template("vencidos.html", resultados=resultados, cantidad=cantidad)



@app.route("/exportar/<int:cliente_id>")
def exportar(cliente_id):
    conn = sqlite3.connect("prestamos.db")
    cursor = conn.cursor()

    cursor.execute("""
        SELECT nombre, apellidos, dni, direccion, telefono, monto, interes, total
        FROM clientes 
        WHERE id = ?
    """, (cliente_id,))
    
    cliente = cursor.fetchone()

    if not cliente:
        conn.close()
        return "Cliente no encontrado"

    cursor.execute("SELECT fecha_pago, cuota, estado FROM cronograma WHERE cliente_id = ?", (cliente_id,))
    cuotas = cursor.fetchall()

    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Cronograma"

    fila = 1

    # 🔹 DATOS CLIENTE
    ws[f"A{fila}"] = "DATOS DEL CLIENTE"
    ws[f"A{fila}"].font = Font(bold=True, size=14)
    fila += 1

    ws.append(["Nombre", cliente[0] + " " + cliente[1]])
    ws.append(["DNI", cliente[2]])
    ws.append(["Dirección", cliente[3]])
    ws.append(["Teléfono", cliente[4]])
    ws.append(["Monto", cliente[5]])
    ws.append(["Interés (%)", cliente[6]])
    ws.append(["Total Préstamo", cliente[7]])

    fila += 2

    # 🔹 CRONOGRAMA
    ws[f"A{fila}"] = "CRONOGRAMA DE PAGOS"
    ws[f"A{fila}"].font = Font(bold=True, size=14)
    fila += 1

    encabezados = ["Fecha Pago", "Monto Cuota", "Estado"]
    ws.append(encabezados)

    for col in range(1, len(encabezados) + 1):
        ws.cell(row=fila, column=col).font = Font(bold=True)

    fila += 1

    for cuota in cuotas:
        ws.append(list(cuota))

    # Ajustar columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(column)].width = max_length + 2

    nombre_archivo = f"cronograma_cliente_{cliente_id}.xlsx"
    wb.save(nombre_archivo)

    return send_file(nombre_archivo, as_attachment=True)






@app.route("/exportar_todos")
def exportar_todos():
    conn = sqlite3.connect("prestamos.db")
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM clientes")
    clientes = cursor.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte General"

    fila = 1

    ws[f"A{fila}"] = "REPORTE GENERAL DE PRÉSTAMOS"
    ws[f"A{fila}"].font = Font(bold=True, size=14)
    fila += 2

    encabezados = [
    "ID",
    "Nombre",
    "DNI",
    "Teléfono",
    "Dirección",
    "Monto",
    "Interés (%)",
    "Total Préstamo",
    "Ganancia (Interés)",
    "Total Pagado",
    "Total Pendiente",
    "Cuotas Pagadas"
]


    ws.append(encabezados)

    for col in range(1, len(encabezados) + 1):
        ws.cell(row=fila, column=col).font = Font(bold=True)

    fila += 1

    for cliente in clientes:
        cliente_id = cliente[0]
        nombre_completo = cliente[1] + " " + cliente[2]
        dni = cliente[3]
        monto = cliente[6]
        total_prestamo = cliente[8]
        telefono = cliente[5]
        direccion = cliente[4]
        interes = cliente[7]


        cursor.execute("SELECT cuota, estado FROM cronograma WHERE cliente_id = ?", (cliente_id,))
        cuotas = cursor.fetchall()

        total_pagado = 0
        cuotas_pagadas = 0

        for cuota in cuotas:
            if cuota[1] == "Pagado":
                total_pagado += cuota[0]
                cuotas_pagadas += 1

        total_pendiente = total_prestamo - total_pagado
        ganancia = total_prestamo - monto

        ws.append([
  	  cliente_id,
 	  nombre_completo,
   	  dni,
	  telefono,
          direccion,
          monto,
          interes,
          total_prestamo,
          ganancia,
          total_pagado,
          total_pendiente,
          cuotas_pagadas
	])


    # Ajustar columnas automáticamente
    for col in ws.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(column)].width = max_length + 2

    conn.close()

    nombre_archivo = "reporte_general.xlsx"
    wb.save(nombre_archivo)

    return send_file(nombre_archivo, as_attachment=True)



import os

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 10000))
    app.run(host="0.0.0.0", port=port)





























